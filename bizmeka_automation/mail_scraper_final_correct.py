#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Bizmeka 메일 스크래퍼 - 최종 정확 버전
실제 li 태그 기반 HTML 구조 사용
"""

import asyncio
from datetime import datetime
import pandas as pd
from playwright.async_api import async_playwright
from cookie_manager import CookieManager
from utils import load_config


async def scrape_bizmeka_mails_final():
    """최종 정확한 메일 스크래핑"""
    
    config = load_config()
    cookie_manager = CookieManager()
    
    cookies = cookie_manager.load_cookies()
    if not cookies:
        print("[ERROR] 쿠키 없음 - manual_login.py 먼저 실행하세요")
        return
    
    print("\n" + "="*60)
    print("Bizmeka 메일 스크래퍼 - 최종 정확 버전")
    print("="*60)
    print("실제 li 태그 구조 기반 추출")
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context()
        await context.add_cookies(cookies)
        page = await context.new_page()
        
        try:
            # 1. 메인 페이지 → 메일 시스템
            print("\n1. 메일 시스템 접속...")
            await page.goto(config['urls']['main'])
            await page.wait_for_timeout(3000)
            
            mail_link = await page.query_selector('a[href*="mail"]')
            if mail_link:
                await mail_link.click()
                await page.wait_for_timeout(3000)
            
            if len(context.pages) > 1:
                page = context.pages[-1]
                print("   → 새 탭 전환")
            
            # 2. 팝업 처리
            print("2. 팝업 처리...")
            for _ in range(3):
                await page.keyboard.press('Escape')
                await page.wait_for_timeout(500)
            
            # 3. 받은메일함 접속
            print("3. 받은메일함 접속...")
            inbox = await page.query_selector('[id^="mnu_Inbox_"]')
            if inbox:
                await inbox.click()
                print("   → 받은메일함 클릭 완료")
                await page.wait_for_timeout(3000)
            
            # 팝업 다시 처리
            for _ in range(2):
                await page.keyboard.press('Escape')
                await page.wait_for_timeout(500)
            
            # 4. 메일 데이터 수집
            print("\n4. 메일 데이터 수집...")
            print("-" * 40)
            
            all_mails = []
            
            for page_num in range(1, 4):  # 3페이지
                print(f"\n페이지 {page_num} 처리 중...")
                
                # 팝업 처리
                await page.keyboard.press('Escape')
                await page.wait_for_timeout(1000)
                
                # li.m_data 요소들 찾기 - 실제 구조
                mail_items = await page.query_selector_all('li.m_data')
                print(f"   → {len(mail_items)}개 메일 아이템 발견")
                
                if len(mail_items) == 0:
                    # 프레임 확인
                    frames = page.frames
                    for frame in frames:
                        if 'mail' in frame.url.lower():
                            print(f"   → 메일 프레임에서 재검색...")
                            mail_items = await frame.query_selector_all('li.m_data')
                            print(f"   → 프레임에서 {len(mail_items)}개 발견")
                            if mail_items:
                                page = frame  # 프레임으로 전환
                                break
                
                page_mails = []
                
                for i, item in enumerate(mail_items):
                    try:
                        # data 속성에서 정보 추출
                        from_name = await item.get_attribute('data-fromname')
                        from_addr = await item.get_attribute('data-fromaddr')
                        mail_id = await item.get_attribute('data-key')
                        
                        # 제목 추출 - p.m_subject 내부
                        subject_elem = await item.query_selector('p.m_subject')
                        subject = ''
                        if subject_elem:
                            subject_text = await subject_elem.inner_text()
                            # 이모지와 불필요한 문자 제거
                            subject = subject_text.strip().replace('🌅', '').replace('&nbsp;', '').strip()
                        
                        # 날짜와 크기 추출 - p.m_info 내부
                        date = ''
                        size = ''
                        info_elem = await item.query_selector('p.m_info')
                        if info_elem:
                            date_elem = await info_elem.query_selector('span.m_date')
                            size_elem = await info_elem.query_selector('span.m_size')
                            
                            if date_elem:
                                date = await date_elem.inner_text()
                            if size_elem:
                                size = await size_elem.inner_text()
                        
                        # 읽음/안읽음 상태
                        is_unread = 'unread' in (await item.get_attribute('class') or '')
                        
                        # 데이터 정리
                        mail_data = {
                            '페이지': page_num,
                            '순번': i + 1,
                            '메일ID': mail_id or '',
                            '보낸사람': (from_name or '').strip(),
                            '이메일주소': (from_addr or '').strip(), 
                            '제목': subject,
                            '날짜': (date or '').strip(),
                            '크기': (size or '').strip(),
                            '읽음상태': '안읽음' if is_unread else '읽음',
                            '수집시간': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        }
                        
                        if mail_data['보낸사람'] or mail_data['제목']:
                            page_mails.append(mail_data)
                            all_mails.append(mail_data)
                            
                            # 처음 3개만 출력
                            if len(page_mails) <= 3:
                                sender = mail_data['보낸사람'][:20] if mail_data['보낸사람'] else 'Unknown'
                                subject = mail_data['제목'][:30] if mail_data['제목'] else 'No subject'
                                print(f"      {len(page_mails)}. {sender}: {subject}...")
                    
                    except Exception as e:
                        print(f"   → 아이템 {i} 처리 오류: {e}")
                        continue
                
                print(f"   → {len(page_mails)}개 메일 수집 완료")
                
                # 다음 페이지로 이동
                if page_num < 3:
                    try:
                        next_page = page_num + 1
                        next_link = await page.query_selector(f'a:has-text("{next_page}")')
                        
                        if next_link:
                            await next_link.click()
                            print(f"   → {next_page}페이지로 이동")
                            await page.wait_for_timeout(3000)
                        else:
                            print(f"   → {next_page}페이지 링크 없음")
                            break
                    except Exception as e:
                        print(f"   → 페이지 이동 실패: {e}")
                        break
            
            # 5. 결과 저장 및 출력
            print("\n" + "="*60)
            if all_mails:
                # DataFrame 생성
                df = pd.DataFrame(all_mails)
                
                # Excel 저장
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'data/bizmeka_mails_final_{timestamp}.xlsx'
                
                # 고급 Excel 저장
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='메일목록')
                    
                    # 워크시트 스타일링
                    worksheet = writer.sheets['메일목록']
                    
                    # 열 너비 최적화
                    column_widths = {
                        'A': 8,   # 페이지
                        'B': 8,   # 순번
                        'C': 25,  # 메일ID
                        'D': 20,  # 보낸사람
                        'E': 30,  # 이메일주소
                        'F': 50,  # 제목
                        'G': 18,  # 날짜
                        'H': 12,  # 크기
                        'I': 10,  # 읽음상태
                        'J': 20   # 수집시간
                    }
                    
                    for col_letter, width in column_widths.items():
                        worksheet.column_dimensions[col_letter].width = width
                    
                    # 헤더 스타일
                    from openpyxl.styles import Font, PatternFill, Alignment
                    header_font = Font(bold=True, size=11)
                    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    header_alignment = Alignment(horizontal='center')
                    
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=1, column=col)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # 안읽은 메일 하이라이트
                    unread_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                    for row_idx, row_data in enumerate(df.itertuples(), start=2):
                        if row_data.읽음상태 == '안읽음':
                            for col in range(1, len(df.columns) + 1):
                                worksheet.cell(row=row_idx, column=col).fill = unread_fill
                
                print(f"[SUCCESS] 메일 수집 대성공! 🎉")
                print(f"\n[수집 결과]")
                print(f"  • 총 메일: {len(all_mails):,}개")
                print(f"  • Excel 파일: {filename}")
                
                # 통계 정보
                page_stats = df.groupby('페이지').size()
                print(f"\n[페이지별 수집량]")
                for page, count in page_stats.items():
                    print(f"  • 페이지 {page}: {count}개")
                
                # 읽음 상태 통계
                status_stats = df['읽음상태'].value_counts()
                print(f"\n[읽음 상태]")
                for status, count in status_stats.items():
                    print(f"  • {status}: {count}개")
                
                # 주요 발신자
                top_senders = df['보낸사람'].value_counts().head(5)
                if len(top_senders) > 0:
                    print(f"\n[주요 발신자 TOP 5]")
                    for sender, count in top_senders.items():
                        if sender:
                            print(f"  • {sender}: {count}개")
                
                # 메일 샘플
                print(f"\n[메일 샘플 - 처음 10개]")
                for i, mail in enumerate(all_mails[:10], 1):
                    status = "📧" if mail['읽음상태'] == '안읽음' else "📖"
                    sender = mail['보낸사람'][:15] if mail['보낸사람'] else 'Unknown'
                    subject = mail['제목'][:35] if mail['제목'] else 'No subject'
                    date = mail['날짜'][:16] if mail['날짜'] else ''
                    print(f"  {i:2d}. {status} [{date}] {sender}: {subject}...")
                
                print(f"\n[추가 정보]")
                print(f"  • 현재 URL: {page.url}")
                print(f"  • 수집 완료 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                
            else:
                print("[WARNING] 수집된 메일이 없습니다")
                
                # 디버깅 자료
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                await page.screenshot(path=f'data/debug_final_{timestamp}.png')
                print(f"[DEBUG] 스크린샷: data/debug_final_{timestamp}.png")
                
                # HTML 소스 확인
                html_content = await page.content()
                with open(f'data/debug_final_{timestamp}.html', 'w', encoding='utf-8') as f:
                    f.write(html_content)
                print(f"[DEBUG] HTML: data/debug_final_{timestamp}.html")
                
                # li.m_data 요소 디버깅
                print(f"[DEBUG] li.m_data 요소 확인...")
                li_elements = await page.query_selector_all('li')
                print(f"   → 총 {len(li_elements)}개 li 요소 발견")
                
                m_data_elements = await page.query_selector_all('.m_data')
                print(f"   → 총 {len(m_data_elements)}개 .m_data 요소 발견")
            
            await browser.close()
            return len(all_mails) if all_mails else 0
            
        except Exception as e:
            print(f"\n[ERROR] 전체 실행 오류: {e}")
            await browser.close()
            return 0


if __name__ == "__main__":
    try:
        result = asyncio.run(scrape_bizmeka_mails_final())
        print(f"\n🏁 프로그램 종료 - 총 {result}개 메일 수집됨")
        
        if result > 0:
            print("\n✅ 메일 스크래핑이 성공적으로 완료되었습니다!")
            print("📊 Excel 파일을 확인해보세요.")
        else:
            print("\n⚠️ 메일을 수집하지 못했습니다.")
            print("🔍 디버그 파일들을 확인해보세요.")
            
    except KeyboardInterrupt:
        print("\n\n⏹️ 사용자에 의해 중단됨")
    except Exception as e:
        print(f"\n❌ 프로그램 오류: {e}")
        import traceback
        traceback.print_exc()