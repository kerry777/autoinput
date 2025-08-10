#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Bizmeka 메일 스크래퍼 - 수동 추출 버전
스크린샷에서 보이는 메일들을 직접 추출
"""

import asyncio
from datetime import datetime
import pandas as pd
from playwright.async_api import async_playwright
from cookie_manager import CookieManager
from utils import load_config


async def extract_visible_mails():
    """보이는 메일 직접 추출"""
    
    config = load_config()
    cookie_manager = CookieManager()
    
    cookies = cookie_manager.load_cookies()
    if not cookies:
        print("[ERROR] 쿠키 없음")
        return
    
    print("\n" + "="*60)
    print("Bizmeka 메일 스크래퍼 - 수동 추출 버전")
    print("="*60)
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context()
        await context.add_cookies(cookies)
        page = await context.new_page()
        
        # 메인 → 메일 → 받은메일함 접속
        print("\n1. 메일 시스템 접속...")
        await page.goto(config['urls']['main'])
        await page.wait_for_timeout(3000)
        
        mail_link = await page.query_selector('a[href*="mail"]')
        if mail_link:
            await mail_link.click()
            await page.wait_for_timeout(3000)
        
        if len(context.pages) > 1:
            page = context.pages[-1]
        
        # 팝업 처리
        for _ in range(3):
            await page.keyboard.press('Escape')
            await page.wait_for_timeout(500)
        
        # 받은메일함 클릭 (동적 선택자)
        inbox = await page.query_selector('[id^="mnu_Inbox_"]')
        if inbox:
            await inbox.click()
            print("2. 받은메일함 접속 완료")
            await page.wait_for_timeout(3000)
        
        # 팝업 다시 처리
        for _ in range(2):
            await page.keyboard.press('Escape')
            await page.wait_for_timeout(500)
        
        print("3. 메일 데이터 수집...")
        all_mails = []
        
        for page_num in range(1, 4):  # 3페이지
            print(f"\n페이지 {page_num} 처리...")
            
            await page.keyboard.press('Escape')
            await page.wait_for_timeout(1000)
            
            # Playwright의 locator 사용 - 더 안전한 방법
            try:
                # 체크박스가 있는 테이블 행 찾기
                checkbox_rows = page.locator('tr:has(input[type="checkbox"]):not(:has(input[id="checkAll"]))')
                count = await checkbox_rows.count()
                
                print(f"   → {count}개 메일 행 발견")
                
                for i in range(count):
                    try:
                        row = checkbox_rows.nth(i)
                        
                        # 각 셀의 텍스트 추출
                        cells = row.locator('td')
                        cell_count = await cells.count()
                        
                        if cell_count >= 6:
                            cell_texts = []
                            for j in range(min(cell_count, 8)):
                                cell_text = await cells.nth(j).inner_text()
                                cell_texts.append(cell_text.strip() if cell_text else '')
                            
                            # 날짜 패턴 확인
                            date_found = False
                            for text in cell_texts:
                                if '2025-' in text or '2024-' in text:
                                    date_found = True
                                    break
                            
                            if date_found:
                                # 메일 데이터 추출
                                # 일반적인 구조: [checkbox][star][attachment][sender][subject][date][size]
                                sender = cell_texts[3] if len(cell_texts) > 3 else ''
                                subject = cell_texts[4] if len(cell_texts) > 4 else ''
                                date = cell_texts[5] if len(cell_texts) > 5 else ''
                                size = cell_texts[6] if len(cell_texts) > 6 else ''
                                
                                # 날짜가 다른 위치에 있을 수 있음
                                if not any('202' in d for d in [date]):
                                    for idx, text in enumerate(cell_texts):
                                        if '202' in text and '-' in text:
                                            date = text
                                            # 다시 추정
                                            if idx >= 2:
                                                sender = cell_texts[idx-2]
                                                subject = cell_texts[idx-1]
                                            if idx < len(cell_texts) - 1:
                                                size = cell_texts[idx+1]
                                            break
                                
                                if sender or subject:
                                    mail_data = {
                                        '페이지': page_num,
                                        '보낸사람': sender,
                                        '제목': subject,
                                        '날짜': date,
                                        '크기': size,
                                        '수집시간': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                    }
                                    all_mails.append(mail_data)
                                    
                                    if len(all_mails) <= 5:
                                        print(f"      • {sender[:20]}: {subject[:30]}...")
                    
                    except Exception as e:
                        continue
                
            except Exception as e:
                print(f"   → 추출 오류: {e}")
            
            # 다음 페이지로 이동
            if page_num < 3:
                try:
                    next_page = page_num + 1
                    next_link = page.locator(f'a:has-text("{next_page}")').first
                    
                    if await next_link.count() > 0:
                        await next_link.click()
                        print(f"   → {next_page}페이지로 이동")
                        await page.wait_for_timeout(2000)
                    else:
                        print(f"   → {next_page}페이지 링크 없음")
                        break
                except:
                    break
        
        # 결과 저장
        print("\n" + "="*60)
        if all_mails:
            df = pd.DataFrame(all_mails)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'data/bizmeka_mails_{timestamp}.xlsx'
            
            # Excel 저장 (스타일링 포함)
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='메일목록')
                
                # 워크시트 스타일링
                worksheet = writer.sheets['메일목록']
                
                # 열 너비 조정
                column_widths = {
                    'A': 10,  # 페이지
                    'B': 25,  # 보낸사람
                    'C': 40,  # 제목
                    'D': 20,  # 날짜
                    'E': 15,  # 크기
                    'F': 20   # 수집시간
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # 헤더 스타일
                from openpyxl.styles import Font, PatternFill
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
                
                for col in range(1, 7):  # A부터 F까지
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
            
            print(f"[SUCCESS] 메일 수집 성공!")
            print(f"\n[수집 결과]")
            print(f"  • 총 메일: {len(all_mails)}개")
            print(f"  • Excel 파일: {filename}")
            
            # 페이지별 통계
            page_stats = df.groupby('페이지').size()
            print(f"\n[페이지별 수집량]")
            for page, count in page_stats.items():
                print(f"  • 페이지 {page}: {count}개")
            
            # 발신자 통계
            if '보낸사람' in df.columns:
                top_senders = df['보낸사람'].value_counts().head(5)
                print(f"\n[주요 발신자]")
                for sender, count in top_senders.items():
                    if sender:
                        print(f"  • {sender}: {count}개")
            
            # 샘플 데이터
            print(f"\n[메일 샘플]")
            for i, mail in enumerate(all_mails[:10], 1):
                sender = mail['보낸사람'][:20] if mail['보낸사람'] else 'Unknown'
                subject = mail['제목'][:35] if mail['제목'] else 'No subject'
                date = mail['날짜'][:16] if mail['날짜'] else ''
                print(f"  {i:2d}. [{date}] {sender}: {subject}...")
            
            print(f"\n[메일함 URL] {page.url}")
            
        else:
            print("[ERROR] 메일을 찾을 수 없습니다")
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            await page.screenshot(path=f'data/error_{timestamp}.png')
        
        await browser.close()
        return len(all_mails) if all_mails else 0


if __name__ == "__main__":
    result = asyncio.run(extract_visible_mails())
    print(f"\n프로그램 종료 - 총 {result}개 메일 수집됨")