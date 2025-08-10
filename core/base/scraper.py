#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
BaseScraper - 모든 사이트 스크래퍼의 기본 클래스
"""

import asyncio
import json
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
from playwright.async_api import Page, BrowserContext
import pandas as pd

from ..utils.popups import PopupHandler
from ..utils.navigation import Navigator
from ..utils.cookies import CookieManager
from ..exceptions.scraping import ScrapingError


class BaseScraper:
    """모든 사이트 스크래퍼의 기본 클래스"""
    
    def __init__(self, site_name: str):
        self.site_name = site_name
        self.site_dir = Path(f"sites/{site_name}")
        self.config_dir = self.site_dir / "config"
        self.data_dir = self.site_dir / "data"
        
        # 유틸리티 초기화
        self.popup_handler = PopupHandler()
        self.navigator = Navigator()
        self.cookie_manager = CookieManager(site_name)
        
        # 상태
        self.page: Optional[Page] = None
        self.context: Optional[BrowserContext] = None
        self.config = self._load_config()
        
    def _load_config(self) -> Dict[str, Any]:
        """사이트별 설정 로드"""
        config_path = self.config_dir / "settings.json"
        if config_path.exists():
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}
    
    def _load_selectors(self) -> Dict[str, str]:
        """사이트별 선택자 로드"""
        selectors_path = self.config_dir / "selectors.json"
        if selectors_path.exists():
            with open(selectors_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}
    
    async def setup_browser(self, context: BrowserContext, page: Page):
        """브라우저 설정"""
        self.context = context
        self.page = page
        
        # 쿠키 로드
        cookies = self.cookie_manager.load_cookies()
        if cookies:
            await context.add_cookies(cookies)
    
    async def close_popups(self) -> bool:
        """팝업 닫기"""
        if not self.page:
            return False
        return await self.popup_handler.close_all(self.page)
    
    async def navigate_to_page(self, page_num: int) -> bool:
        """페이지 이동"""
        if not self.page:
            return False
        return await self.navigator.go_to_page(self.page, page_num)
    
    async def wait_and_retry(self, operation, max_attempts: int = 3, delay: float = 1.0):
        """재시도 로직"""
        for attempt in range(max_attempts):
            try:
                return await operation()
            except Exception as e:
                if attempt == max_attempts - 1:
                    raise ScrapingError(f"Max attempts reached: {e}")
                await asyncio.sleep(delay * (attempt + 1))
    
    def save_data_to_excel(self, data: List[Dict], filename: str = None) -> str:
        """Excel 저장"""
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{self.site_name}_{timestamp}.xlsx"
        
        filepath = self.data_dir / filename
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        df = pd.DataFrame(data)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='데이터')
            
            # 스타일링
            worksheet = writer.sheets['데이터']
            
            # 헤더 스타일
            from openpyxl.styles import Font, PatternFill
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
        
        return str(filepath)
    
    def log(self, message: str, level: str = "INFO"):
        """로깅"""
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f"[{timestamp}] [{level}] [{self.site_name.upper()}] {message}")
    
    # 추상 메서드 - 하위 클래스에서 구현
    async def scrape(self) -> List[Dict[str, Any]]:
        """스크래핑 메인 로직 - 하위 클래스에서 구현"""
        raise NotImplementedError("scrape method must be implemented by subclass")
    
    async def login(self) -> bool:
        """로그인 로직 - 하위 클래스에서 구현"""
        raise NotImplementedError("login method must be implemented by subclass")