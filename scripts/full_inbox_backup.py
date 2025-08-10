#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Bizmeka 받은메일함 전체 백업
21페이지 모든 메일을 완벽한 Excel로 추출
"""

import asyncio
import sys
from pathlib import Path
from datetime import datetime

# 프로젝트 루트를 Python path에 추가
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from playwright.async_api import async_playwright
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class FullInboxBackup:
    """전체 받은메일함 백업 클래스"""
    
    def __init__(self):
        self.all_mails = []
        self.page_stats = {}
        self.start_time = datetime.now()
        
    async def setup_browser_and_login(self):
        """브라우저 설정 및 로그인"""
        self.p = async_playwright()
        playwright = await self.p.__aenter__()
        
        browser = await playwright.chromium.launch(headless=False)
        context = await browser.new_context(
            locale='ko-KR',
            timezone_id='Asia/Seoul'
        )
        page = await context.new_page()
        
        # 쿠키 로드
        cookie_file = Path("sites/bizmeka/data/cookies.json")
        if cookie_file.exists():
            import json
            with open(cookie_file, 'r', encoding='utf-8') as f:
                cookies = json.load(f)
            await context.add_cookies(cookies)
            print("[OK] 쿠키 로드 완료")
        else:
            print("[ERROR] 쿠키 파일이 없습니다. manual_login.py를 먼저 실행하세요")
            await browser.close()
            return None, None, None
        
        # 메인 페이지 → 메일 시스템
        await page.goto("https://bizmeka.com")
        await page.wait_for_timeout(3000)
        
        mail_link = await page.query_selector('a[href*="mail"]')
        if mail_link:
            await mail_link.click()
            await page.wait_for_timeout(3000)
        
        # 새 탭 전환
        if len(context.pages) > 1:
            page = context.pages[-1]
        
        # 받은메일함 접속
        await self.close_popups(page)
        
        inbox = await page.query_selector('[id^="mnu_Inbox_"]')
        if inbox:
            await inbox.click()
            print("[OK] 받은메일함 접속")
            await page.wait_for_timeout(3000)
        
        await self.close_popups(page)
        
        return browser, context, page
    
    async def close_popups(self, page):
        """팝업 닫기"""
        for _ in range(3):
            await page.keyboard.press('Escape')
            await page.wait_for_timeout(500)
    
    async def extract_page_mails(self, page, page_num):
        """페이지별 메일 추출"""
        await self.close_popups(page)
        
        print(f"페이지 {page_num} 추출 중...", end=" ")
        
        page_mails = []
        
        try:
            # 프레임 확인
            target_page = page
            for frame in page.frames:
                if 'mail' in frame.url.lower():
                    target_page = frame
                    break
            
            # 메일 아이템 추출
            mail_items = await target_page.query_selector_all('li.m_data')
            
            for i, item in enumerate(mail_items):
                try:
                    # 기본 속성
                    mail_id = await item.get_attribute('data-key') or ''
                    from_name = await item.get_attribute('data-fromname') or ''
                    from_addr = await item.get_attribute('data-fromaddr') or ''
                    item_class = await item.get_attribute('class') or ''
                    
                    # 읽음 상태
                    is_unread = 'unread' in item_class
                    
                    # 제목 추출 및 정리
                    subject = ''
                    subject_elem = await item.query_selector('p.m_subject')
                    if subject_elem:
                        subject_raw = await subject_elem.inner_text()
                        # 이모지와 특수문자 정리
                        subject = self.clean_subject(subject_raw)
                    
                    # 날짜 추출
                    date = ''
                    date_elem = await item.query_selector('span.m_date')
                    if date_elem:
                        date = await date_elem.inner_text()
                    
                    # 크기 추출
                    size = ''
                    size_elem = await item.query_selector('span.m_size')
                    if size_elem:
                        size = await size_elem.inner_text()
                    
                    # 중요도/별표
                    star_important = False
                    star_elem = await item.query_selector('button.btn_star')
                    if star_elem:
                        star_attr = await star_elem.get_attribute('star')
                        star_important = star_attr == '1'
                    
                    # 첨부파일 여부
                    has_attachment = False
                    attach_elem = await item.query_selector('.m_file')
                    if attach_elem:
                        attach_content = await attach_elem.inner_text()
                        has_attachment = bool(attach_content and attach_content.strip())
                    
                    # 메일 데이터 구성
                    mail_data = {
                        '페이지': page_num,
                        '순번': len(self.all_mails) + 1,
                        '메일ID': mail_id,
                        '보낸사람': from_name.strip(),
                        '보낸사람_이메일': from_addr.strip(),
                        '제목': subject.strip(),
                        '수신일시': date.strip(),
                        '크기': size.strip(),
                        '읽음상태': '미읽음' if is_unread else '읽음',
                        '중요표시': '★' if star_important else '',
                        '첨부파일': '📎' if has_attachment else '',
                        '수집시간': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    
                    if mail_data['보낸사람'] or mail_data['제목']:
                        page_mails.append(mail_data)
                        self.all_mails.append(mail_data)
                
                except Exception as e:
                    continue
            
            print(f"→ {len(page_mails)}개 추출")
            self.page_stats[page_num] = len(page_mails)
            
        except Exception as e:
            print(f"→ 오류: {e}")
            self.page_stats[page_num] = 0
        
        return len(page_mails)
    
    def clean_subject(self, raw_subject):
        """제목 정리"""
        if not raw_subject:
            return ''
        
        subject = raw_subject.strip()
        
        # 이모지 제거
        subject = subject.replace('🌅', '').replace('📧', '').replace('🔔', '')
        subject = subject.replace('⚡', '').replace('🎯', '').replace('💡', '')
        
        # HTML 엔티티 처리
        subject = subject.replace('&nbsp;', ' ')
        subject = subject.replace('&amp;', '&')
        subject = subject.replace('&lt;', '<')
        subject = subject.replace('&gt;', '>')
        
        # 개행 및 탭 정리
        subject = subject.replace('\n', ' ').replace('\t', ' ')
        
        # 다중 공백 정리
        while '  ' in subject:
            subject = subject.replace('  ', ' ')
        
        return subject.strip()
    
    async def navigate_to_page(self, page, target_page):
        """페이지 이동"""
        try:
            await self.close_popups(page)
            
            # 페이지 번호 링크 찾기
            selectors = [
                f'a:has-text("{target_page}")',
                f'[onclick*="page={target_page}"]',
                f'.pagination a:has-text("{target_page}")'
            ]
            
            for selector in selectors:
                try:
                    page_link = await page.query_selector(selector)
                    if page_link:
                        await page_link.click()
                        await page.wait_for_timeout(3000)
                        return True
                except:
                    continue
            
            # 다음 버튼 시도 (페이지 번호가 안보일 때)
            next_btn = await page.query_selector('a:has-text("다음"), a[title="다음"]')
            if next_btn:
                await next_btn.click()
                await page.wait_for_timeout(3000)
                return True
            
            return False
            
        except Exception as e:
            print(f"페이지 이동 오류: {e}")
            return False
    
    def save_to_excel(self, filename=None):
        """고급 Excel 저장"""
        if not self.all_mails:
            print("[ERROR] 저장할 데이터가 없습니다")
            return None
        
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'sites/bizmeka/data/inbox_full_backup_{timestamp}.xlsx'
        
        # DataFrame 생성
        df = pd.DataFrame(self.all_mails)
        
        # 파일 경로 생성
        filepath = Path(filename)
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        # 고급 Excel 저장
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # 메인 데이터 시트
            df.to_excel(writer, sheet_name='받은메일함_전체', index=False)
            
            # 통계 시트 생성
            stats_data = {
                '항목': ['총 메일 수', '총 페이지 수', '미읽음 메일', '읽음 메일', '중요 표시', '첨부파일 있음', '수집 시작', '수집 완료', '소요 시간'],
                '값': [
                    len(self.all_mails),
                    len(self.page_stats),
                    len(df[df['읽음상태'] == '미읽음']),
                    len(df[df['읽음상태'] == '읽음']),
                    len(df[df['중요표시'] == '★']),
                    len(df[df['첨부파일'] == '📎']),
                    self.start_time.strftime('%Y-%m-%d %H:%M:%S'),
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    str(datetime.now() - self.start_time).split('.')[0]
                ]
            }
            stats_df = pd.DataFrame(stats_data)
            stats_df.to_excel(writer, sheet_name='수집통계', index=False)
            
            # 페이지별 통계 시트
            page_data = {
                '페이지': list(self.page_stats.keys()),
                '메일수': list(self.page_stats.values())
            }
            page_df = pd.DataFrame(page_data)
            page_df.to_excel(writer, sheet_name='페이지별통계', index=False)
            
            # 스타일링 적용
            self.apply_excel_styling(writer, df)
        
        return str(filepath)
    
    def apply_excel_styling(self, writer, df):
        """Excel 스타일링"""
        workbook = writer.book
        worksheet = writer.sheets['받은메일함_전체']
        
        # 헤더 스타일
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 테두리
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 헤더 스타일 적용
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 열 너비 자동 조정
        column_widths = {
            'A': 8,   # 페이지
            'B': 8,   # 순번
            'C': 25,  # 메일ID
            'D': 20,  # 보낸사람
            'E': 30,  # 보낸사람_이메일
            'F': 50,  # 제목
            'G': 18,  # 수신일시
            'H': 12,  # 크기
            'I': 10,  # 읽음상태
            'J': 8,   # 중요표시
            'K': 8,   # 첨부파일
            'L': 20   # 수집시간
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # 미읽음 메일 하이라이트
        unread_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        important_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        for row_idx in range(2, len(df) + 2):  # 헤더 제외하고 데이터 행
            mail_data = df.iloc[row_idx - 2]
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col)
                cell.border = thin_border
                
                # 미읽음 메일 하이라이트
                if mail_data['읽음상태'] == '미읽음':
                    cell.fill = unread_fill
                
                # 중요 메일 하이라이트
                elif mail_data['중요표시'] == '★':
                    cell.fill = important_fill
    
    async def run_full_backup(self, max_pages=21):
        """전체 백업 실행"""
        print("="*60)
        print("Bizmeka 받은메일함 전체 백업")
        print("="*60)
        print(f"목표: {max_pages}페이지 전체 추출")
        print(f"시작 시간: {self.start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        print("-"*60)
        
        browser, context, page = await self.setup_browser_and_login()
        if not browser:
            return
        
        try:
            # 페이지별 추출
            for page_num in range(1, max_pages + 1):
                # 첫 페이지가 아니면 페이지 이동
                if page_num > 1:
                    if not await self.navigate_to_page(page, page_num):
                        print(f"페이지 {page_num} 이동 실패 - {page_num-1}페이지까지만 수집")
                        break
                
                # 메일 추출
                count = await self.extract_page_mails(page, page_num)
                
                # 진행상황 표시
                if page_num % 5 == 0:
                    total_so_far = len(self.all_mails)
                    elapsed = datetime.now() - self.start_time
                    print(f"--- {page_num}페이지 완료 | 현재까지 {total_so_far}개 메일 | 소요시간: {elapsed} ---")
            
            # 결과 저장
            print("\n" + "="*60)
            print("수집 완료! Excel 파일 생성 중...")
            
            filepath = self.save_to_excel()
            
            if filepath:
                # 최종 결과
                end_time = datetime.now()
                elapsed_time = end_time - self.start_time
                
                print(f"🎉 전체 백업 완료!")
                print(f"\n📊 수집 결과:")
                print(f"  • 총 메일: {len(self.all_mails):,}개")
                print(f"  • 수집 페이지: {len(self.page_stats)}페이지")
                print(f"  • 소요 시간: {elapsed_time}")
                print(f"  • Excel 파일: {filepath}")
                
                # 페이지별 통계
                print(f"\n📄 페이지별 수집량:")
                total_pages = len(self.page_stats)
                for i, (page_num, count) in enumerate(self.page_stats.items(), 1):
                    if i <= 10 or i > total_pages - 5:  # 처음 10개와 마지막 5개만 표시
                        print(f"  페이지 {page_num}: {count}개")
                    elif i == 11:
                        print(f"  ... (중간 {total_pages - 15}개 페이지 생략)")
                
                # 메일 상태 통계
                if self.all_mails:
                    df = pd.DataFrame(self.all_mails)
                    
                    print(f"\n📧 메일 상태:")
                    if '읽음상태' in df.columns:
                        status_counts = df['읽음상태'].value_counts()
                        for status, count in status_counts.items():
                            print(f"  • {status}: {count:,}개")
                    
                    if '중요표시' in df.columns:
                        important_count = len(df[df['중요표시'] == '★'])
                        if important_count > 0:
                            print(f"  • 중요 메일: {important_count:,}개")
                    
                    if '첨부파일' in df.columns:
                        attachment_count = len(df[df['첨부파일'] == '📎'])
                        if attachment_count > 0:
                            print(f"  • 첨부파일 있음: {attachment_count:,}개")
                
                print(f"\n💾 파일 위치: {filepath}")
            
        except Exception as e:
            print(f"백업 오류: {e}")
            import traceback
            traceback.print_exc()
            
        finally:
            await browser.close()
            await self.p.__aexit__(None, None, None)


async def main():
    """메인 실행 함수"""
    backup = FullInboxBackup()
    await backup.run_full_backup(max_pages=21)


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n사용자에 의해 중단됨")
    except Exception as e:
        print(f"프로그램 오류: {e}")