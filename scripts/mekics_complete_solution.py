#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
MEK-ICS 완전한 솔루션 - 새 로그인부터 다운로드까지
"""

import asyncio
import json
import csv
from pathlib import Path
from datetime import datetime
from playwright.async_api import async_playwright
import time


async def complete_mekics_download():
    """완전한 MEK-ICS 다운로드 자동화"""
    
    # 경로 설정
    site_dir = Path("sites/mekics")
    data_dir = site_dir / "data"
    downloads_dir = data_dir / "downloads"
    downloads_dir.mkdir(parents=True, exist_ok=True)
    
    # 설정 파일 로드
    config_path = site_dir / "config" / "settings.json"
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
    
    async with async_playwright() as p:
        # 브라우저 시작 (headless=False로 화면 표시)
        browser = await p.chromium.launch(
            headless=False,
            args=['--start-maximized']
        )
        
        # 컨텍스트 생성
        context = await browser.new_context(
            locale='ko-KR',
            timezone_id='Asia/Seoul',
            viewport={'width': 1920, 'height': 1080},
            accept_downloads=True
        )
        
        # 페이지 생성
        page = await context.new_page()
        
        try:
            print("\n" + "="*60)
            print("MEK-ICS 완전 자동화 솔루션")
            print("="*60)
            
            # ========== 1단계: 로그인 ==========
            print("\n[1단계] MEK-ICS 로그인")
            print("-"*40)
            
            # 로그인 페이지 접속
            print("  로그인 페이지 접속...")
            await page.goto("https://it.mek-ics.com/mekics/login/login.do")
            await page.wait_for_timeout(3000)
            
            # 로그인 수행
            print(f"  아이디 입력: {config['credentials']['username']}")
            await page.fill('input[name="USER_ID"]', config['credentials']['username'])
            await page.wait_for_timeout(500)
            
            print("  비밀번호 입력...")
            await page.fill('input[name="USER_PWD"]', config['credentials']['password'])
            await page.wait_for_timeout(500)
            
            print("  로그인 버튼 클릭...")
            await page.click('button.btn_login')
            
            # 로그인 완료 대기
            print("  로그인 처리 중...")
            await page.wait_for_timeout(5000)
            
            # 2FA 처리 (필요한 경우)
            try:
                # 2FA 입력 필드 확인
                auth_input = await page.wait_for_selector('input[type="text"]', timeout=3000)
                if auth_input:
                    print("\n  [2FA 필요] 화면에서 직접 입력해주세요.")
                    print("  30초 대기 중...")
                    await page.wait_for_timeout(30000)
            except:
                print("  2FA 없음 - 계속 진행")
            
            # 메인 페이지 확인
            await page.wait_for_timeout(3000)
            current_url = page.url
            
            if "main" in current_url or "index" in current_url:
                print("  ✓ 로그인 성공!")
                
                # 쿠키 저장
                cookies = await context.cookies()
                cookie_file = data_dir / "cookies.json"
                with open(cookie_file, 'w', encoding='utf-8') as f:
                    json.dump(cookies, f, ensure_ascii=False, indent=2)
                print("  ✓ 쿠키 저장 완료")
            else:
                print("  ! 로그인 확인 필요")
            
            # ========== 2단계: 매출현황조회 접속 ==========
            print("\n[2단계] 매출현황조회 페이지 접속")
            print("-"*40)
            
            print("  매출현황조회 페이지로 이동...")
            await page.goto("https://it.mek-ics.com/mekics/sales/ssa450skrv.do?authoUser=A")
            await page.wait_for_timeout(5000)
            
            # ExtJS 로드 확인
            extjs_loaded = await page.evaluate("() => typeof Ext !== 'undefined'")
            if extjs_loaded:
                print("  ✓ ExtJS 로드 완료")
            
            # ========== 3단계: 조회 조건 설정 ==========
            print("\n[3단계] 조회 조건 설정")
            print("-"*40)
            
            # 날짜 설정
            start_date = "2021-01-01"
            end_date = datetime.now().strftime('%Y-%m-%d')
            print(f"  날짜 범위: {start_date} ~ {end_date}")
            
            # 조건 설정 스크립트
            setup_result = await page.evaluate("""
                () => {
                    try {
                        // LOT표시 '아니오' 선택
                        const noRadio = Ext.getCmp('radiofield-1078');
                        if(noRadio) {
                            noRadio.setValue(true);
                            console.log('LOT표시: 아니오');
                        }
                        
                        // 국내/해외 '전체' 선택
                        const allRadio = Ext.getCmp('radiofield-1081');
                        if(allRadio) {
                            allRadio.setValue(true);
                            console.log('국내/해외: 전체');
                        }
                        
                        // 날짜 설정
                        const dateFields = Ext.ComponentQuery.query('datefield');
                        if(dateFields.length >= 2) {
                            // 시작일: 2021년 1월 1일
                            dateFields[0].setValue(new Date(2021, 0, 1));
                            // 종료일: 오늘
                            dateFields[1].setValue(new Date());
                            console.log('날짜 설정 완료');
                        }
                        
                        return true;
                    } catch(e) {
                        return false;
                    }
                }
            """)
            
            if setup_result:
                print("  ✓ 조건 설정 완료")
            
            # ========== 4단계: 데이터 조회 ==========
            print("\n[4단계] 데이터 조회 실행")
            print("-"*40)
            
            print("  F2 키 입력 (조회 실행)...")
            await page.keyboard.press('F2')
            
            # 조회 대기
            print("  데이터 로딩 중...")
            for i in range(15):
                await page.wait_for_timeout(1000)
                
                # 데이터 로드 확인
                status = await page.evaluate("""
                    () => {
                        const grids = Ext.ComponentQuery.query('grid');
                        if(grids.length > 0) {
                            const store = grids[0].getStore();
                            return {
                                count: store.getCount(),
                                loading: store.isLoading ? store.isLoading() : false
                            };
                        }
                        return {count: 0, loading: true};
                    }
                """)
                
                if status['count'] > 0 and not status['loading']:
                    print(f"  ✓ 데이터 로드 완료: {status['count']:,}건")
                    break
                
                if i % 3 == 2:
                    print(f"    {i+1}초 경과...")
            
            # CSV 팝업 확인
            print("\n  CSV 팝업 확인...")
            csv_popup = None
            try:
                csv_popup = await page.wait_for_selector('text="CSV 형태로 다운로드"', timeout=3000)
            except:
                pass
            
            if csv_popup:
                print("  ✓ CSV 다운로드 팝업 감지 (대용량 데이터)")
                
                # 다운로드 처리
                print("\n[5단계] CSV 다운로드")
                print("-"*40)
                
                # 다운로드 이벤트 준비
                async with page.expect_download() as download_info:
                    # '예' 버튼 클릭
                    yes_button = await page.wait_for_selector('span.x-btn-inner:has-text("예")')
                    print("  '예' 버튼 클릭...")
                    await yes_button.click()
                    
                    # 다운로드 대기
                    print("  다운로드 대기 중...")
                    download = await download_info.value
                    
                    # 파일 저장
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"매출현황_전체_{timestamp}.csv"
                    save_path = downloads_dir / filename
                    
                    await download.save_as(str(save_path))
                    
                    file_size = save_path.stat().st_size / (1024 * 1024)
                    
                    # 데이터 건수 확인
                    with open(save_path, 'r', encoding='utf-8-sig') as f:
                        line_count = sum(1 for line in f) - 1
                    
                    print(f"\n  ✅ CSV 다운로드 성공!")
                    print(f"     파일명: {filename}")
                    print(f"     크기: {file_size:.2f} MB")
                    print(f"     데이터: {line_count:,}건")
                    print(f"     경로: {save_path}")
            
            else:
                # Store 데이터 추출
                print("\n[5단계] Store 데이터 추출")
                print("-"*40)
                
                # 데이터 추출
                extracted = await page.evaluate("""
                    () => {
                        const grids = Ext.ComponentQuery.query('grid');
                        if(grids.length === 0) return null;
                        
                        const grid = grids[0];
                        const store = grid.getStore();
                        const columns = grid.getColumns();
                        
                        // 헤더 추출
                        const headers = [];
                        const dataIndexes = [];
                        
                        columns.forEach(col => {
                            if(!col.hidden && col.dataIndex) {
                                headers.push(col.text || col.dataIndex);
                                dataIndexes.push(col.dataIndex);
                            }
                        });
                        
                        // 데이터 추출
                        const rows = [];
                        store.each(record => {
                            const row = [];
                            dataIndexes.forEach(dataIndex => {
                                let value = record.get(dataIndex);
                                if(value === null || value === undefined) {
                                    value = '';
                                }
                                row.push(String(value));
                            });
                            rows.push(row);
                        });
                        
                        return {
                            headers: headers,
                            rows: rows,
                            count: rows.length
                        };
                    }
                """)
                
                if extracted and extracted['count'] > 0:
                    # CSV 저장
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"매출현황_store_{timestamp}.csv"
                    csv_path = downloads_dir / filename
                    
                    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
                        writer = csv.writer(f)
                        writer.writerow(extracted['headers'])
                        writer.writerows(extracted['rows'])
                    
                    print(f"\n  ✅ Store 데이터 추출 성공!")
                    print(f"     파일명: {filename}")
                    print(f"     데이터: {extracted['count']:,}건")
                    print(f"     경로: {csv_path}")
                    
                    # Excel 변환
                    try:
                        import openpyxl
                        
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "매출현황"
                        
                        # 헤더 추가
                        for col_idx, header in enumerate(extracted['headers'], 1):
                            ws.cell(row=1, column=col_idx, value=header)
                        
                        # 데이터 추가
                        for row_idx, row_data in enumerate(extracted['rows'], 2):
                            for col_idx, value in enumerate(row_data, 1):
                                ws.cell(row=row_idx, column=col_idx, value=value)
                        
                        # Excel 저장
                        excel_filename = filename.replace('.csv', '.xlsx')
                        excel_path = downloads_dir / excel_filename
                        wb.save(excel_path)
                        
                        print(f"     Excel: {excel_filename}")
                        
                    except ImportError:
                        print("     (Excel 변환 생략 - openpyxl 필요)")
            
            # ========== 완료 ==========
            print("\n" + "="*60)
            print("✅ 작업 완료!")
            print(f"저장 위치: {downloads_dir}")
            print("="*60)
            
            # 스크린샷
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            screenshot = data_dir / f"complete_{timestamp}.png"
            await page.screenshot(path=str(screenshot), full_page=True)
            print(f"\n스크린샷: {screenshot}")
            
        except Exception as e:
            print(f"\n❌ 오류 발생: {e}")
            import traceback
            traceback.print_exc()
            
            # 오류 스크린샷
            error_screenshot = data_dir / f"error_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            await page.screenshot(path=str(error_screenshot))
            print(f"오류 스크린샷: {error_screenshot}")
            
        finally:
            print("\n브라우저를 30초 후 종료합니다...")
            await page.wait_for_timeout(30000)
            await browser.close()


if __name__ == "__main__":
    print("MEK-ICS 완전 자동화 시작...")
    asyncio.run(complete_mekics_download())