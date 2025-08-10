#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
MEK-ICS 최종 솔루션 - Store 데이터 추출 방식 (검증됨)
"""

import asyncio
import json
import sys
from pathlib import Path
from datetime import datetime
import csv

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from playwright.async_api import async_playwright


async def extract_sales_data():
    site_dir = Path("sites/mekics")
    data_dir = site_dir / "data"
    downloads_dir = data_dir / "downloads"
    downloads_dir.mkdir(parents=True, exist_ok=True)
    
    config_path = site_dir / "config" / "settings.json"
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        
        context = await browser.new_context(
            locale='ko-KR',
            timezone_id='Asia/Seoul',
            viewport={'width': 1920, 'height': 1080}
        )
        
        # 쿠키 로드
        cookie_file = data_dir / "cookies.json"
        if cookie_file.exists():
            with open(cookie_file, 'r', encoding='utf-8') as f:
                cookies = json.load(f)
            await context.add_cookies(cookies)
        
        page = await context.new_page()
        
        
        try:
            print("\nMEK-ICS 매출 데이터 추출 (검증된 방법)")
            print("="*60)
            
            # 페이지 접속
            print("\n[1] 매출현황조회 페이지 접속...")
            await page.goto("https://it.mek-ics.com/mekics/sales/ssa450skrv.do?authoUser=A")
            await page.wait_for_timeout(5000)
            
            # 조회 조건 설정
            print("\n[2] 조회 조건 설정...")
            
            # 날짜 설정 (자동)
            start_date = "20250801"  # 8월 1일
            end_date = "20250810"    # 8월 10일
            print(f"  날짜 설정: {start_date} ~ {end_date}")
            
            await page.evaluate(f"""
                () => {{
                    // LOT표시 '아니오'
                    const noRadio = Ext.getCmp('radiofield-1078');
                    if(noRadio) noRadio.setValue(true);
                    
                    // 국내/해외 '전체'
                    const allRadio = Ext.getCmp('radiofield-1081');
                    if(allRadio) allRadio.setValue(true);
                    
                    // 날짜 설정
                    const dateFields = Ext.ComponentQuery.query('datefield');
                    if(dateFields.length >= 2) {{
                        // 날짜 형식 변환
                        const startYear = parseInt('{start_date}'.substr(0, 4));
                        const startMonth = parseInt('{start_date}'.substr(4, 2)) - 1;
                        const startDay = parseInt('{start_date}'.substr(6, 2));
                        
                        const endYear = parseInt('{end_date}'.substr(0, 4));
                        const endMonth = parseInt('{end_date}'.substr(4, 2)) - 1;
                        const endDay = parseInt('{end_date}'.substr(6, 2));
                        
                        dateFields[0].setValue(new Date(startYear, startMonth, startDay));
                        dateFields[1].setValue(new Date(endYear, endMonth, endDay));
                    }}
                }}
            """)
            
            print(f"  설정 완료: {start_date} ~ {end_date}")
            
            # 조회 실행
            print("\n[3] 데이터 조회 실행 (F2)...")
            await page.keyboard.press('F2')
            
            # 데이터 로딩 대기
            print("  데이터 로딩 중...")
            await page.wait_for_timeout(10000)
            
            # CSV 팝업 처리
            try:
                csv_popup = await page.wait_for_selector('text="CSV 형태로 다운로드"', timeout=2000)
                if csv_popup:
                    print("  CSV 팝업 감지 - '예' 클릭")
                    yes_button = await page.wait_for_selector('span.x-btn-inner:has-text("예")', timeout=1000)
                    if yes_button:
                        await yes_button.click()
                        await page.wait_for_timeout(2000)
            except:
                pass  # 팝업이 없으면 계속 진행
            
            # 데이터 개수 확인
            data_count = await page.evaluate("""
                () => {
                    const grids = Ext.ComponentQuery.query('grid');
                    if(grids.length > 0) {
                        return grids[0].getStore().getCount();
                    }
                    return 0;
                }
            """)
            
            print(f"  조회 완료: {data_count:,}건")
            
            if data_count == 0:
                print("\n경고: 조회된 데이터가 없습니다.")
                return
            
            # Store 데이터 추출
            print("\n[4] 데이터 추출 중...")
            
            extracted_data = await page.evaluate("""
                () => {
                    const grids = Ext.ComponentQuery.query('grid');
                    if(grids.length === 0) return null;
                    
                    const grid = grids[0];
                    const store = grid.getStore();
                    const columns = grid.getColumns();
                    
                    // 컬럼 헤더 추출 (숨겨진 컬럼 제외)
                    const headers = [];
                    const dataIndexes = [];
                    
                    columns.forEach(col => {
                        if(!col.hidden && col.dataIndex) {
                            headers.push(col.text || col.dataIndex);
                            dataIndexes.push(col.dataIndex);
                        }
                    });
                    
                    // 데이터 추출
                    const rows = [];
                    store.each(record => {
                        const row = [];
                        dataIndexes.forEach(dataIndex => {
                            let value = record.get(dataIndex);
                            if(value === null || value === undefined) {
                                value = '';
                            }
                            row.push(String(value));
                        });
                        rows.push(row);
                    });
                    
                    return {
                        headers: headers,
                        rows: rows,
                        count: rows.length
                    };
                }
            """)
            
            if not extracted_data:
                print("오류: 데이터 추출 실패")
                return
            
            print(f"  추출 완료: {extracted_data['count']:,}건")
            
            # CSV 파일 저장
            print("\n[5] 파일 저장...")
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"매출현황_{start_date}_{end_date}_{timestamp}"
            
            # CSV 저장
            csv_file = downloads_dir / f"{filename}.csv"
            with open(csv_file, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(extracted_data['headers'])
                writer.writerows(extracted_data['rows'])
            
            print(f"  CSV 저장: {csv_file}")
            
            # Excel 변환 옵션
            try:
                import openpyxl
                print("\n[6] Excel 변환...")
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "매출현황"
                
                # 헤더 추가
                for col_idx, header in enumerate(extracted_data['headers'], 1):
                    ws.cell(row=1, column=col_idx, value=header)
                
                # 데이터 추가
                for row_idx, row_data in enumerate(extracted_data['rows'], 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 열 너비 자동 조정
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                excel_file = downloads_dir / f"{filename}.xlsx"
                wb.save(excel_file)
                print(f"  Excel 저장: {excel_file}")
                
            except ImportError:
                print("\n  (openpyxl 미설치로 Excel 변환 생략)")
            
            # 스크린샷
            screenshot = data_dir / f"extract_{timestamp}.png"
            await page.screenshot(path=str(screenshot))
            print(f"\n스크린샷: {screenshot}")
            
            print("\n" + "="*60)
            print("작업 완료!")
            print(f"저장 위치: {downloads_dir}")
            
        except Exception as e:
            print(f"\n오류: {e}")
            import traceback
            traceback.print_exc()
        
        finally:
            print("\n30초 후 종료...")
            await page.wait_for_timeout(30000)
            await browser.close()


if __name__ == "__main__":
    asyncio.run(extract_sales_data())