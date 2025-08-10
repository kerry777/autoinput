#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
MEK-ICS 작동 확인된 솔루션 - Store 데이터 추출
이미 검증된 방법: CSV와 Excel 파일 생성 성공
"""

import asyncio
import json
import csv
from pathlib import Path
from datetime import datetime
from playwright.async_api import async_playwright


async def login_mekics(page, config):
    """MEK-ICS 로그인"""
    print("\n[로그인 필요] MEK-ICS 로그인...")
    await page.goto("https://it.mek-ics.com/mekics/login/login.do")
    await page.wait_for_timeout(2000)
    
    # 아이디 입력
    await page.fill('input[name="USER_ID"]', config['credentials']['username'])
    await page.wait_for_timeout(500)
    
    # 비밀번호 입력
    await page.fill('input[name="USER_PWD"]', config['credentials']['password'])
    await page.wait_for_timeout(500)
    
    # 로그인 버튼 클릭
    await page.click('button.btn_login')
    await page.wait_for_timeout(5000)
    
    # 쿠키 저장
    cookies = await page.context.cookies()
    cookie_file = Path("sites/mekics/data/cookies.json")
    with open(cookie_file, 'w', encoding='utf-8') as f:
        json.dump(cookies, f, ensure_ascii=False, indent=2)
    
    print("  로그인 완료, 쿠키 저장됨")


async def extract_sales_data():
    """검증된 Store 데이터 추출 방식"""
    site_dir = Path("sites/mekics")
    data_dir = site_dir / "data"
    downloads_dir = data_dir / "downloads"
    downloads_dir.mkdir(parents=True, exist_ok=True)
    
    config_path = site_dir / "config" / "settings.json"
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        
        context = await browser.new_context(
            locale='ko-KR',
            timezone_id='Asia/Seoul',
            viewport={'width': 1920, 'height': 1080}
        )
        
        page = await context.new_page()
        
        try:
            print("\nMEK-ICS 매출 데이터 추출 (작동 확인된 방법)")
            print("="*60)
            
            # 로그인 시도
            await login_mekics(page, config)
            
            # 매출현황조회 페이지 접속
            print("\n[1] 매출현황조회 페이지 접속...")
            await page.goto("https://it.mek-ics.com/mekics/sales/ssa450skrv.do?authoUser=A")
            await page.wait_for_timeout(5000)
            
            # 조회 조건 설정
            print("\n[2] 조회 조건 설정...")
            
            # 대용량 데이터 테스트
            start_date = "20210101"
            end_date = datetime.now().strftime('%Y%m%d')
            print(f"  날짜: {start_date} ~ {end_date}")
            
            await page.evaluate(f"""
                () => {{
                    // LOT표시 '아니오'
                    const noRadio = Ext.getCmp('radiofield-1078');
                    if(noRadio) noRadio.setValue(true);
                    
                    // 국내/해외 '전체'
                    const allRadio = Ext.getCmp('radiofield-1081');
                    if(allRadio) allRadio.setValue(true);
                    
                    // 날짜 설정
                    const dateFields = Ext.ComponentQuery.query('datefield');
                    if(dateFields.length >= 2) {{
                        dateFields[0].setValue(new Date(2021, 0, 1));
                        dateFields[1].setValue(new Date());
                    }}
                }}
            """)
            
            # 조회 실행
            print("\n[3] 데이터 조회 실행 (F2)...")
            await page.keyboard.press('F2')
            
            # CSV 팝업 처리 (대용량 시)
            print("  팝업 확인 중...")
            csv_download_started = False
            
            try:
                # CSV 팝업 대기
                csv_popup = await page.wait_for_selector('text="CSV 형태로 다운로드"', timeout=10000)
                if csv_popup:
                    print("\n[대용량 데이터 감지] CSV 다운로드 팝업 표시")
                    
                    # 다운로드 준비
                    async with page.expect_download() as download_info:
                        # '예' 버튼 클릭
                        yes_button = await page.wait_for_selector('span.x-btn-inner:has-text("예")', timeout=2000)
                        if yes_button:
                            print("  '예' 버튼 클릭 -> 다운로드 시작")
                            await yes_button.click()
                            
                            try:
                                # 다운로드 대기 (최대 60초)
                                download = await download_info.value
                                
                                # 파일 저장
                                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                                save_path = downloads_dir / f"sales_large_{timestamp}.csv"
                                await download.save_as(str(save_path))
                                
                                file_size = save_path.stat().st_size / (1024 * 1024)
                                
                                # 데이터 건수 확인
                                with open(save_path, 'r', encoding='utf-8-sig') as f:
                                    line_count = sum(1 for line in f) - 1
                                
                                print("\n" + "="*60)
                                print("[성공] 대용량 CSV 다운로드 완료!")
                                print(f"파일: {save_path.name}")
                                print(f"크기: {file_size:.2f} MB")
                                print(f"데이터: {line_count:,}건")
                                print(f"경로: {save_path}")
                                print("="*60)
                                
                                csv_download_started = True
                                
                            except Exception as download_error:
                                print(f"  다운로드 실패: {download_error}")
                                
            except:
                print("  CSV 팝업 없음 - 일반 조회")
            
            # CSV 다운로드가 실패했거나 팝업이 없는 경우
            if not csv_download_started:
                await page.wait_for_timeout(10000)
                
                # Store 데이터 확인
                data_count = await page.evaluate("""
                    () => {
                        const grids = Ext.ComponentQuery.query('grid');
                        if(grids.length > 0) {
                            return grids[0].getStore().getCount();
                        }
                        return 0;
                    }
                """)
                
                print(f"\n[Store 데이터] {data_count:,}건 로드됨")
                
                if data_count > 0:
                    print("\n[4] Store 데이터 추출...")
                    
                    # 데이터 추출
                    extracted_data = await page.evaluate("""
                        () => {
                            const grids = Ext.ComponentQuery.query('grid');
                            if(grids.length === 0) return null;
                            
                            const grid = grids[0];
                            const store = grid.getStore();
                            const columns = grid.getColumns();
                            
                            // 헤더 추출
                            const headers = [];
                            const dataIndexes = [];
                            
                            columns.forEach(col => {
                                if(!col.hidden && col.dataIndex) {
                                    headers.push(col.text || col.dataIndex);
                                    dataIndexes.push(col.dataIndex);
                                }
                            });
                            
                            // 데이터 추출
                            const rows = [];
                            store.each(record => {
                                const row = [];
                                dataIndexes.forEach(dataIndex => {
                                    let value = record.get(dataIndex);
                                    if(value === null || value === undefined) {
                                        value = '';
                                    }
                                    row.push(String(value));
                                });
                                rows.push(row);
                            });
                            
                            return {
                                headers: headers,
                                rows: rows,
                                count: rows.length
                            };
                        }
                    """)
                    
                    if extracted_data:
                        print(f"  추출 완료: {extracted_data['count']:,}건")
                        
                        # CSV 저장
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        csv_file = downloads_dir / f"sales_store_{timestamp}.csv"
                        
                        with open(csv_file, 'w', encoding='utf-8-sig', newline='') as f:
                            writer = csv.writer(f)
                            writer.writerow(extracted_data['headers'])
                            writer.writerows(extracted_data['rows'])
                        
                        print(f"\n[완료] Store 데이터 CSV 저장")
                        print(f"  파일: {csv_file}")
                        print(f"  데이터: {extracted_data['count']:,}건")
                        
                        # Excel 변환 시도
                        try:
                            import openpyxl
                            wb = openpyxl.Workbook()
                            ws = wb.active
                            ws.title = "매출현황"
                            
                            # 헤더
                            for col_idx, header in enumerate(extracted_data['headers'], 1):
                                ws.cell(row=1, column=col_idx, value=header)
                            
                            # 데이터
                            for row_idx, row_data in enumerate(extracted_data['rows'], 2):
                                for col_idx, value in enumerate(row_data, 1):
                                    ws.cell(row=row_idx, column=col_idx, value=value)
                            
                            excel_file = downloads_dir / f"sales_store_{timestamp}.xlsx"
                            wb.save(excel_file)
                            print(f"  Excel: {excel_file}")
                        except:
                            pass
            
            # 스크린샷
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            screenshot = data_dir / f"working_{timestamp}.png"
            await page.screenshot(path=str(screenshot))
            print(f"\n스크린샷: {screenshot}")
            
            print("\n" + "="*60)
            print("작업 완료!")
            print(f"저장 위치: {downloads_dir}")
            
        except Exception as e:
            print(f"\n오류: {e}")
            import traceback
            traceback.print_exc()
        
        finally:
            print("\n30초 후 종료...")
            await page.wait_for_timeout(30000)
            await browser.close()


if __name__ == "__main__":
    asyncio.run(extract_sales_data())